import io
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


VERDE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
AMARILLO = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ROJO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ENCABEZADO = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ENCABEZADO_FUENTE = Font(color="FFFFFF", bold=True)


def _generar_excel(preview: list[dict], reporte: dict) -> bytes:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resultados de Evaluación"
    headers1 = [
        "Competencia ID", "Nombre", "Nivel", "Etiqueta",
        "Justificación", "Citas", "Secciones Fuente",
        "Estado Revisión", "Estado Final", "Confianza", "JPC",
    ]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = ENCABEZADO
        cell.font = ENCABEZADO_FUENTE
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for row_i, r in enumerate(preview, 2):
        cid = r["competencia_id"]
        jpc = None
        for tr in reporte.get("trazabilidad_competencias", []):
            if tr.get("competencia_id") == cid:
                jpc = tr.get("JPC")
                break
        values = [
            cid,
            r.get("competencia_nombre", ""),
            r.get("nivel", 0),
            r.get("nivel_label", ""),
            r.get("justificacion", ""),
            "\n".join(r.get("citas", [])),
            ", ".join(r.get("secciones_fuente", [])),
            r.get("estado_revision", ""),
            r.get("estado_final", "pendiente"),
            f'{r.get("confianza", 0):.1%}',
            f'{jpc:.1%}' if jpc is not None else "N/A",
        ]
        for col, val in enumerate(values, 1):
            cell = ws1.cell(row=row_i, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border
        estado_revision = r.get("estado_revision", "")
        fill = None
        if estado_revision == "respaldo_suficiente":
            fill = VERDE
        elif estado_revision == "requiere_revision":
            fill = AMARILLO
        elif estado_revision == "sin_evidencia":
            fill = ROJO
        if fill:
            for col in range(1, len(headers1) + 1):
                ws1.cell(row=row_i, column=col).fill = fill
    for col in range(1, len(headers1) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 25
    ws1.column_dimensions["E"].width = 50
    ws1.column_dimensions["F"].width = 50
    ws2 = wb.create_sheet("Reporte de Procesamiento")
    trazabilidad = reporte.get("trazabilidad_competencias", [])
    jpc_fields = [
        "competencia_id", "estado_cobertura", "estado_final",
        "nivel_asignado", "chunks_recuperados",
        "C_cobertura_citas", "S_pertinencia_seccion",
        "R_similitud_promedio", "F_confianza", "JPC", "JPC_aplicable",
    ]
    for col, h in enumerate(jpc_fields, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = ENCABEZADO
        cell.font = ENCABEZADO_FUENTE
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for row_i, tr in enumerate(trazabilidad, 2):
        for col, field in enumerate(jpc_fields, 1):
            val = tr.get(field, "")
            if isinstance(val, float):
                val = round(val, 4)
            cell = ws2.cell(row=row_i, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border
    tiempos_row = len(trazabilidad) + 3
    ws2.cell(row=tiempos_row, column=1, value="TIEMPOS").font = Font(bold=True)
    tiempos = reporte.get("tiempos", {})
    for col, key in enumerate(
        ["T_procesamiento_automatico_min", "T_revision_humana_min", "T_ajustes_min", "T_IA_total_min"], 2
    ):
        val = tiempos.get(key)
        ws2.cell(row=tiempos_row, column=col, value=key.replace("_", " "))
        ws2.cell(row=tiempos_row + 1, column=col, value=val if val is not None else "N/A")
        ws2.cell(row=tiempos_row, column=col).font = Font(bold=True)
    ajustes = reporte.get("historial_ajustes", [])
    if ajustes:
        ajustes_row = tiempos_row + 3
        ws2.cell(row=ajustes_row, column=1, value="HISTORIAL DE AJUSTES").font = Font(bold=True)
        aj_headers = ["ajuste_id", "competencia_id", "solicitud_usuario", "capas_reprocesadas", "duracion_min"]
        for col, h in enumerate(aj_headers, 1):
            ws2.cell(row=ajustes_row + 1, column=col, value=h).font = Font(bold=True)
        for row_i, aj in enumerate(ajustes, ajustes_row + 2):
            for col, field in enumerate(aj_headers, 1):
                val = aj.get(field, "")
                if isinstance(val, list):
                    val = ", ".join(val)
                ws2.cell(row=row_i, column=col, value=val)
    for col in range(1, len(jpc_fields) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 22
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def render():
    st.title("Resultados y Exportación")

    if "pipeline_state" not in st.session_state:
        st.warning("No hay resultados disponibles. Ejecuta el pipeline primero.")
        return

    state = st.session_state["pipeline_state"]
    c7 = state["c7"]
    preview = c7["vista_preliminar"]["resultados_competencias"]
    reporte = c7["reporte_procesamiento"]

    st.subheader("Vista Preliminar de Resultados")
    for r in preview:
        cid = r["competencia_id"]
        estado = r.get("estado_final", "pendiente")
        if estado == "aprobada":
            st.success(f"{cid}: Nivel {r['nivel']} - {r['nivel_label']} (Aprobada)")
        elif estado == "rechazada":
            st.error(f"{cid}: Nivel {r['nivel']} - {r['nivel_label']} (Rechazada)")
        elif estado == "modificada":
            st.info(f"{cid}: Nivel {r['nivel']} - {r['nivel_label']} (Modificada)")
        else:
            st.warning(f"{cid}: Nivel {r['nivel']} - {r['nivel_label']} (Pendiente)")

    st.divider()
    st.subheader("Reporte de Procesamiento")
    trazabilidad = reporte.get("trazabilidad_competencias", [])
    if trazabilidad:
        df_traz = pd.DataFrame(trazabilidad)
        st.dataframe(df_traz, use_container_width=True)

    tiempos = reporte.get("tiempos", {})
    if tiempos:
        st.subheader("Tiempos del Pipeline")
        tc1, tc2, tc3, tc4 = st.columns(4)
        tc1.metric("Procesamiento Automático", f'{tiempos.get("T_procesamiento_automatico_min", "N/A")} min')
        tc2.metric("Revisión Humana", f'{tiempos.get("T_revision_humana_min", "N/A")} min')
        tc3.metric("Ajustes", f'{tiempos.get("T_ajustes_min", "N/A")} min')
        tc4.metric("Total IA", f'{tiempos.get("T_IA_total_min", "N/A")} min')

    historial = reporte.get("historial_ajustes", [])
    if historial:
        st.subheader("Historial de Ajustes")
        df_hist = pd.DataFrame(historial)
        st.dataframe(df_hist, use_container_width=True)

    st.divider()
    st.subheader("Exportar")

    if st.button("Generar y Descargar Excel", type="primary", use_container_width=True):
        try:
            excel_bytes = _generar_excel(preview, reporte)
            st.download_button(
                label="Descargar .xlsx",
                data=excel_bytes,
                file_name="evaluacion_competencias.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.success("Archivo Excel generado correctamente.")
        except Exception as e:
            st.error(f"Error al generar el Excel: {e}")

    col1, col2 = st.columns(2)
    with col1:
        if st.button("Volver a Pipeline", use_container_width=True):
            st.session_state["page"] = "pipeline"
            st.rerun()
    with col2:
        if st.button("Nueva Evaluación", use_container_width=True):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()
