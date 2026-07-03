import streamlit as st
import base64
import io
from html import escape
from pipeline.cohorts import get_cohort, compute_cohort_macro, update_cohort_name
from pipeline.cohorts import remove_report_from_cohort
from pipeline.persistence import load_report
from pipeline.reportes_export import build_export_index, exportar_excel_multi_hoja
from openpyxl import load_workbook
from ui.components import page_hero, badge, action_tiles
from ui.icons import chart, upload, download, trash


def _formatear_tipo(tipo: str) -> str:
    nombre = tipo.replace("_", " ").title()
    nombre = nombre.replace("Pre ", "Pre-")
    nombre = nombre.replace("Practica", "Práctica")
    return nombre


def _cohort_profile_summary_html(macro: dict) -> str:
    competencias = macro.get("competencias", {})
    comp_list = sorted(competencias.items(), key=lambda item: item[1].get("tasa_aprobacion", 0))
    total = len(comp_list)
    aprobadas = sum(1 for _, data in comp_list if data.get("tasa_aprobacion", 0) >= 0.70)
    perfil_pct = aprobadas / total * 100 if total else 0
    debiles = [(cid, data) for cid, data in comp_list if data.get("tasa_aprobacion", 0) < 0.70]

    if debiles:
        debiles_html = "".join(
            f'<span><b>{escape(cid)}</b> {data.get("tasa_aprobacion", 0) * 100:.0f}%</span>'
            for cid, data in debiles[:8]
        )
    else:
        debiles_html = '<span><b>Sin brechas</b> 100%</span>'

    return (
        '<div class="cohort-profile-summary">'
        '<div class="cohort-profile-card primary">'
        '<span>Perfil de egreso aprobado</span>'
        f'<strong>{perfil_pct:.0f}%</strong>'
        f'<p>{aprobadas} de {total} competencias evidenciadas con al menos 70% de aprobación.</p>'
        '</div>'
        '<div class="cohort-profile-card weak">'
        '<span>Competencias débiles de esta cohorte</span>'
        f'<strong>{len(debiles)}</strong>'
        f'<div class="cohort-weak-list">{debiles_html}</div>'
        '</div>'
        '</div>'
    )


@st.dialog("Eliminar Informes")
def delete_reports_dialog(cohort_id, cohort_name):
    cohort = get_cohort(cohort_id)
    if not cohort:
        st.warning("No se encontró la cohorte.")
        return

    report_ids = cohort.get("report_ids", [])
    if not report_ids:
        st.info("No hay informes para eliminar.")
        if st.button("Cerrar", use_container_width=True):
            st.rerun()
        return

    st.write(f"Selecciona los informes de **'{cohort_name}'** que deseas eliminar:")

    selected_rids = []
    for rid in report_ids:
        report = load_report(rid)
        name = report.pdf_name if report else rid[:8]
        if st.checkbox(name, key=f"del_check_{rid}"):
            selected_rids.append(rid)

    st.markdown("")
    col1, col2 = st.columns(2)
    with col1:
        if st.button("Cancelar", use_container_width=True, key="cancel_delete_dialog"):
            st.session_state.pop("_delete_reports_open", None)
            st.rerun()
    with col2:
        if st.button("Eliminar seleccionados", type="primary", use_container_width=True, key="confirm_delete_dialog_btn"):
            if not selected_rids:
                st.warning("Selecciona al menos un informe para eliminar.")
            else:
                for rid in selected_rids:
                    remove_report_from_cohort(cohort_id, rid)
                st.session_state.pop("_delete_reports_open", None)
                st.session_state["report_count"] = max(0, st.session_state.get("report_count", 0) - len(selected_rids))
                st.rerun()


def render():
    cohort_id = st.session_state.get("selected_cohort_id")
    cohort = get_cohort(cohort_id) if cohort_id else None

    if not cohort:
        st.warning("No se encontró la cohorte seleccionada.")
        return

    tipo_label = _formatear_tipo(cohort.get("tipo_documento", ""))
    n_reports = len(cohort.get("report_ids", []))
    macro = compute_cohort_macro(cohort_id)

    meta_items = [
        badge(tipo_label, "outline"),
        f"{n_reports} informe{'s' if n_reports != 1 else ''}",
    ]
    if cohort.get("created_at"):
        meta_items.append(f'Creada: {cohort["created_at"][:10]}')

    page_hero(
        "Configuración: " + cohort["name"],
        subtitle="Acciones rápidas y gestión de la cohorte.",
        meta_items=meta_items,
        back_target="cohorts",
    )

    st.markdown(_cohort_profile_summary_html(macro), unsafe_allow_html=True)

    st.subheader("Acciones Rápidas")
    
    export_index = build_export_index(cohort.get("report_ids", []))
    excel_full = exportar_excel_multi_hoja(export_index)

    excel_full.seek(0)
    wb = load_workbook(excel_full)
    if "Reporte de Procesamiento" in wb.sheetnames:
        del wb["Reporte de Procesamiento"]
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    excel_resultados_b64 = base64.b64encode(out.getvalue()).decode()

    excel_full.seek(0)
    wb2 = load_workbook(excel_full)
    keep = "Reporte de Procesamiento"
    for sn in list(wb2.sheetnames):
        if sn != keep:
            del wb2[sn]
    out2 = io.BytesIO()
    wb2.save(out2)
    out2.seek(0)
    excel_procesamiento_b64 = base64.b64encode(out2.getvalue()).decode()

    col_tiles = st.columns(5)
    with col_tiles[0]:
        action_tiles([{
            "icon": chart(28, 28, "currentColor"),
            "title": "Resultados Macro",
            "desc": "Ver resumen agregado de la cohorte",
            "tone": "tone-red",
            "url": f"?page=cohort_macro&cid={cohort_id}"
        }])
        
    with col_tiles[1]:
        action_tiles([{
            "icon": upload(28, 28, "currentColor"),
            "title": "Agregar Informes",
            "desc": "Subir más informes a esta cohorte",
            "tone": "tone-blue",
            "url": f"?page=upload&cid={cohort_id}&new=0"
        }])

    with col_tiles[2]:
        action_tiles([{
            "icon": download(28, 28, "currentColor"),
            "title": "Exportar Resultados",
            "desc": "Resumen macro y evaluación",
            "tone": "tone-yellow",
            "download": excel_resultados_b64,
            "filename": f"{cohort['name']}_resultados.xlsx"
        }])

    with col_tiles[3]:
        action_tiles([{
            "icon": download(28, 28, "currentColor"),
            "title": "Descargar Reporte Tecnico",
            "desc": "Solo datos técnicos",
            "tone": "tone-yellow",
            "download": excel_procesamiento_b64,
            "filename": f"{cohort['name']}_procesamiento.xlsx"
        }])

    with col_tiles[4]:
        action_tiles([{
            "icon": trash(28, 28, "currentColor"),
            "title": "Eliminar Informes",
            "desc": "Eliminar informes de la cohorte",
            "danger": True,
            "tone": "tone-danger",
            "url": f"?page=cohort_config&action=delete_reports&cid={cohort_id}"
        }])

    if st.session_state.get("_action") == "delete_reports":
        st.session_state.pop("_action", None)
        st.session_state["_delete_reports_open"] = True

    if st.session_state.get("_delete_reports_open"):
        delete_reports_dialog(cohort_id, cohort["name"])

    st.subheader("Configuración")
    with st.container(border=True):
        st.markdown('<div class="uandes-form-section">Edición de datos</div>', unsafe_allow_html=True)
        col_name, col_meta = st.columns([2, 1])
        with col_name:
            new_name = st.text_input("Nombre de la cohorte", value=cohort["name"], key="edit_cohort_name")
            if new_name != cohort["name"]:
                if update_cohort_name(cohort_id, new_name):
                    st.success("Nombre actualizado.")
                    st.rerun()
        with col_meta:
             st.markdown(f'''
             <div class="config-meta-card">
               <p><strong>Tipo:</strong> {tipo_label}</p>
               <p><strong>Creada:</strong> {cohort.get("created_at", "")[:10]}</p>
             </div>
             ''', unsafe_allow_html=True)
