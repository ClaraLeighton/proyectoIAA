import streamlit as st
from html import escape
from textwrap import dedent

from pipeline.cohorts import get_cohort, compute_cohort_macro
from pipeline.persistence import load_report
from pipeline.reportes_export import build_export_index, exportar_excel_multi_hoja
from openpyxl import load_workbook
import io
from ui.components import page_hero, empty_state, badge


LEVEL_COLORS = {"0": "#ef4444", "1": "#f97316", "2": "#2e9cdb", "3": "#22c55e"}
LEVEL_LABELS = {"0": "Sin evidencia", "1": "Solo teoría", "2": "Uso concreto", "3": "Dominio técnico"}
LEVEL_SHORT = {"0": "SE", "1": "TE", "2": "UC", "3": "DT"}


def _promedio_a_nivel(promedio: float) -> str:
    if promedio >= 2.5: return "3"
    if promedio >= 1.5: return "2"
    if promedio >= 0.5: return "1"
    return "0"

def _promedio_a_label(promedio: float) -> str:
    return LEVEL_LABELS[_promedio_a_nivel(promedio)]


def _hex_to_rgba(h: str, a: float) -> str:
    h = h.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return f"rgba({r},{g},{b},{a})"


def _formatear_tipo(tipo: str) -> str:
    nombre = tipo.replace("_", " ").title()
    nombre = nombre.replace("Pre ", "Pre-")
    nombre = nombre.replace("Practica", "Práctica")
    return nombre


def _sort_competencia_id(cid: str):
    digits = "".join(ch for ch in cid if ch.isdigit())
    return int(digits) if digits else cid


def _clamp_pct(value: float) -> float:
    return max(0, min(100, value))


def _estado_competencia(tasa_aprobacion: float) -> tuple[str, str, str]:
    if tasa_aprobacion >= 0.70:
        return "Evidenciada", "ok", "La mayoría de los informes demuestra esta competencia de forma consistente (≥ 70% alcanza grado 2 o superior)."
    if tasa_aprobacion >= 0.50:
        return "Parcialmente evidenciada", "mid", "La competencia aparece en varios informes, pero aún no de forma consistente (50–69% alcanza grado 2 o superior)."
    return "Escasa evidencia", "risk", "Existe poca o ninguna evidencia de esta competencia en los informes evaluados (< 50% alcanza grado 2 o superior)."


def _clasificar_competencia(tasa_aprobacion: float) -> str:
    if tasa_aprobacion >= 0.70:
        return "Evidenciada"
    if tasa_aprobacion >= 0.50:
        return "Parcialmente evidenciada"
    return "Escasa evidencia"


def _clasif_css_class(tasa_aprobacion: float) -> str:
    c = _clasificar_competencia(tasa_aprobacion)
    return "evidenciada" if c == "Evidenciada" else "parcial" if "Parcialmente" in c else "escasa"


def _generate_macro_only_excel(report_ids: list[str]) -> io.BytesIO:
    data = exportar_excel_multi_hoja(build_export_index(report_ids))
    wb = load_workbook(data)
    if "Reporte de Procesamiento" in wb.sheetnames:
        del wb["Reporte de Procesamiento"]
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _generate_processing_excel(cohort_name: str, report_ids: list[str]) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Procesamiento"

    headers = [
        "Informe", "Tiempo procesamiento (min)", "Modelo embeddings",
        "Proveedor", "Cantidad chunks", "Cantidad embeddings",
        "Competencias evaluadas", "Secciones detectadas",
        "Secciones ausentes", "Logs disponibles",
    ]
    ws.append(headers)
    hf = PatternFill("solid", fgColor="E7E5E4")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = hf
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for rid in report_ids:
        report = load_report(rid)
        if not report or report.estado != "completado":
            continue

        ps = report.pipeline_state
        rp = report.reporte_procesamiento
        tiempos = rp.get("tiempos", {})
        trazas = rp.get("trazabilidad_competencias", [])
        c2 = ps.get("c2", {})
        c3 = ps.get("c3", {})
        c4 = ps.get("c4", {})

        ws.append([
            report.pdf_name,
            tiempos.get("T_procesamiento_automatico_min", ""),
            c4.get("reporte", {}).get("modelo_embeddings", ""),
            c4.get("reporte", {}).get("proveedor", ""),
            c3.get("reporte", {}).get("total_chunks", ""),
            c4.get("reporte", {}).get("chunks_embeddings", ""),
            len(trazas),
            ", ".join(c2.get("secciones_detectadas", [])),
            ", ".join(c2.get("secciones_ausentes", [])),
            f"Ajustes: {len(rp.get('historial_ajustes', []))}" if rp.get("historial_ajustes") else "Sin ajustes",
        ])

    ws.freeze_panes = "A2"
    for col, w in enumerate(["A:44", "B:24", "C:32", "D:18", "E:18", "F:20", "G:22", "H:44", "I:44", "J:30"], 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else None].width = int(w.split(":")[1])  # simplified
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 44
    ws.column_dimensions["I"].width = 44
    ws.column_dimensions["J"].width = 30
    wb.save(output)
    output.seek(0)
    return output


def _build_stacked_bar(dist: dict, total: int) -> str:
    if total <= 0:
        return '<div class="macro-stack empty"></div>'
    segments = []
    for lvl in ["0", "1", "2", "3"]:
        count = dist.get(lvl, 0)
        pct = count / total * 100
        if count:
            segments.append(
                f'<span title="{LEVEL_LABELS[lvl]}: {count}" '
                f'style="width:{pct:.1f}%;background:{LEVEL_COLORS[lvl]}">{count}</span>'
            )
    return f'<div class="macro-stack">{"".join(segments)}</div>'


def _legend_html(nivel_dist: dict, total: int) -> str:
    items = []
    for lvl in ["0", "1", "2", "3"]:
        count = nivel_dist.get(lvl, 0)
        pct = count / total * 100 if total else 0
        items.append(
            f'<span title="Del total de evaluaciones en los informes: {count} de {total} con grado {lvl} ({LEVEL_LABELS[lvl]})"><i style="background:{LEVEL_COLORS[lvl]}"></i>'
            f'{LEVEL_LABELS[lvl]} <strong>{count}</strong> ({pct:.0f}%)</span>'
        )
    return f'<div class="macro-legend">{"".join(items)}</div>'


def _build_line_chart(comp_list: list[tuple[str, dict]]) -> str:
    if not comp_list:
        return '<div class="macro-line-empty">Sin datos para graficar.</div>'

    width = 680
    height = 260
    pad_x = 42
    pad_top = 28
    pad_bottom = 46
    plot_w = width - pad_x * 2
    plot_h = height - pad_top - pad_bottom
    denom = max(1, len(comp_list) - 1)

    def points_for(metric: str) -> list[tuple[float, float, str, float]]:
        pts = []
        for idx, (cid, data) in enumerate(comp_list):
            value = data.get(metric, 0) * 100
            x = pad_x + (idx / denom) * plot_w
            y = pad_top + (1 - value / 100) * plot_h
            pts.append((x, y, cid, value))
        return pts

    score_pts = points_for("score_pct")
    aprob_pts = points_for("tasa_aprobacion")
    score_attr = " ".join(f"{x:.1f},{y:.1f}" for x, y, _, _ in score_pts)
    aprob_attr = " ".join(f"{x:.1f},{y:.1f}" for x, y, _, _ in aprob_pts)
    area_attr = f"{pad_x},{height - pad_bottom} {score_attr} {width - pad_x},{height - pad_bottom}"

    score_dots = []
    aprob_dots = []
    labels = []
    for x, y, cid, value in score_pts:
        score_dots.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="4.5"><title>{escape(cid)} grado de evidencia: {value:.0f}%</title></circle>')
        labels.append(f'<text x="{x:.1f}" y="{height - 19}" text-anchor="middle">{escape(cid)}</text>')
    for x, y, cid, value in aprob_pts:
        aprob_dots.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="3.8"><title>{escape(cid)} tasa de logro: {value:.0f}%</title></circle>')

    grid = []
    for pct in [0, 25, 50, 75, 100]:
        y = pad_top + (1 - pct / 100) * plot_h
        grid.append(
            f'<line x1="{pad_x}" y1="{y:.1f}" x2="{width - pad_x}" y2="{y:.1f}"></line>'
            f'<text x="10" y="{y + 4:.1f}">{pct}%</text>'
        )

    return (
        '<div class="macro-line-chart">'
        f'<svg viewBox="0 0 {width} {height}" role="img" aria-label="Frecuencia y profundidad por competencia">'
        '<defs><linearGradient id="macroScoreFill" x1="0" x2="0" y1="0" y2="1">'
        '<stop offset="0%" stop-color="#CE0019" stop-opacity="0.24"></stop>'
        '<stop offset="100%" stop-color="#CE0019" stop-opacity="0.02"></stop>'
        '</linearGradient></defs>'
        f'<g class="macro-chart-grid-lines">{"".join(grid)}</g>'
        f'<polygon class="macro-area" points="{area_attr}"></polygon>'
        f'<polyline class="macro-score-line" points="{score_attr}"></polyline>'
        f'<polyline class="macro-aprob-line" points="{aprob_attr}"></polyline>'
        f'<g class="macro-score-dots">{"".join(score_dots)}</g>'
        f'<g class="macro-aprob-dots">{"".join(aprob_dots)}</g>'
        f'<g class="macro-chart-labels">{"".join(labels)}</g>'
        '</svg>'
        '</div>'
    )


def _build_bar_chart(comp_list: list[tuple[str, dict]]) -> str:
    if not comp_list:
        return '<div class="macro-line-empty">Sin datos para graficar.</div>'

    ordered = sorted(comp_list, key=lambda item: _sort_competencia_id(item[0]))
    max_items = ordered
    max_value = max((data.get("tasa_aprobacion", 0) for _, data in max_items), default=1)
    max_value = max(max_value, 0.01)
    bars = []
    for cid, data in max_items:
        aprob = data.get("tasa_aprobacion", 0) * 100
        estado, estado_cls, _ = _estado_competencia(data.get("tasa_aprobacion", 0))
        height = 22 + (data.get("tasa_aprobacion", 0) / max_value) * 92
        bars.append(
            f'<div class="macro-mini-bar {estado_cls}" title="{escape(cid)} · {estado} · {aprob:.0f}% de logro (grado ≥ 2)">'
            f'<i style="height:{height:.1f}px"></i><strong>{aprob:.0f}%</strong><span>{escape(cid)}</span></div>'
        )
    return f'<div class="macro-mini-bars">{"".join(bars)}</div>'


def _html_block(template: str) -> str:
    return "".join(line.strip() for line in dedent(template).splitlines())


def _macro_dashboard_html(cohort, tipo_label, g, nivel_dist, total_comps, competencias):
    aprob_pct = _clamp_pct(g["tasa_aprobacion_global"] * 100)
    nivel_pct = _clamp_pct((g["nivel_promedio_global"] / 3) * 100 if g["nivel_promedio_global"] else 0)
    comp_list = sorted(competencias.items(), key=lambda item: _sort_competencia_id(item[0]))
    cumplidas = sum(1 for _, c in comp_list if c.get("tasa_aprobacion", 0) >= 0.70)
    desarrollo = sum(1 for _, c in comp_list if 0.50 <= c.get("tasa_aprobacion", 0) < 0.70)
    brecha = sum(1 for _, c in comp_list if c.get("tasa_aprobacion", 0) < 0.50)
    perfil_pct = cumplidas / len(comp_list) * 100 if comp_list else 0

    total_aprobadas = nivel_dist.get("2", 0) + nivel_dist.get("3", 0)

    sorted_by_score = sorted(comp_list, key=lambda item: item[1].get("score_pct", 0), reverse=True)
    line_chart = _build_line_chart(comp_list)
    bar_chart = _build_bar_chart(comp_list)
    top_rows = []
    for cid, data in sorted_by_score[:4]:
        pct = data.get("score_pct", 0) * 100
        top_rows.append(
            f'<div class="macro-rank-row">'
            f'<strong>{escape(cid)}</strong>'
            f'<span>{escape(data.get("nombre", "")[:52])}</span>'
            f'<em title="{data.get("aprobadas", 0)} de {data.get("total_reportes", 0)} reportes con grado ≥ 2">{pct:.0f}%</em>'
            f'</div>'
        )

    weak_rows = []
    sorted_by_approval = sorted(comp_list, key=lambda item: item[1].get("tasa_aprobacion", 0))
    for cid, data in sorted_by_approval[:3]:
        pct = data.get("tasa_aprobacion", 0) * 100
        weak_rows.append(
            f'<div class="macro-gap-row">'
            f'<div><strong>{escape(cid)}</strong><span>{escape(data.get("nombre", "")[:48])}</span></div>'
            f'<em title="{data.get("aprobadas", 0)} de {data.get("total_reportes", 0)} reportes con grado ≥ 2">{pct:.0f}% logro</em>'
            f'</div>'
        )

    tiles = []
    for cid, data in comp_list:
        aprob = data.get("tasa_aprobacion", 0) * 100
        estado, estado_cls, _ = _estado_competencia(data.get("tasa_aprobacion", 0))
        tip_logro = f'{data.get("aprobadas", 0)} de {data.get("total_reportes", 0)} reportes alcanzaron grado 2 (Uso concreto) o superior en esta competencia'
        tiles.append(
            f'<div class="macro-comp-tile {estado_cls}">'
            f'<div class="macro-comp-top"><strong>{escape(cid)}</strong><span>{estado}</span></div>'
            f'<p>{escape(data.get("nombre", ""))}</p>'
            f'<div class="macro-comp-meter"><i style="width:{_clamp_pct(aprob):.1f}%"></i></div>'
            f'<div class="macro-comp-foot"><span title="{tip_logro}">{aprob:.0f}% aprobación</span></div>'
            f'</div>'
        )

    matrix_rows = []
    for cid, data in comp_list:
        dist = data.get("distribucion", {})
        total = data.get("total_reportes", 0)
        aprob = data.get("tasa_aprobacion", 0) * 100
        cells = []
        for lvl in ["0", "1", "2", "3"]:
            count = dist.get(lvl, 0)
            heat = 1.0 if count > 0 else 0.15
            cells.append(
                f'<td><span class="macro-heat-cell" '
                f'style="--heat:{heat:.2f};background:{LEVEL_COLORS[lvl]}">'
                f'{count}</span></td>'
            )
        n_prom = data.get("nivel_promedio", 0)
        n_key = _promedio_a_nivel(n_prom)
        matrix_rows.append(
            f'<tr>'
            f'<td><strong>{escape(cid)}</strong><small>{escape(data.get("nombre", ""))}</small></td>'
            f'<td title="Promedio {n_prom:.2f}/3 — {data.get("score_actual", 0)}/{data.get("score_max", 0)} puntos en {data.get("total_reportes", 0)} evaluaciones"><i style="background:{LEVEL_COLORS[n_key]};width:9px;height:9px;border-radius:999px;display:inline-block;margin-right:5px"></i>{escape(_promedio_a_label(n_prom))}</td>'
            f'<td><div class="macro-table-meter"><i style="width:{_clamp_pct(aprob):.1f}%"></i></div><b title="{data.get("aprobadas", 0)} de {data.get("total_reportes", 0)} reportes con grado ≥ 2">{aprob:.0f}%</b></td>'
            f'{"".join(cells)}'
            f'<td><span class="macro-clasif {_clasif_css_class(data.get("tasa_aprobacion", 0))}">{_clasificar_competencia(data.get("tasa_aprobacion", 0))}</span></td>'
            f'</tr>'
        )

    return _html_block(f"""
    <div class="macro-dashboard">
      <section class="macro-hero-panel">
        <div class="macro-hero-copy">
          <div class="macro-eyebrow">{escape(tipo_label)} · Cohorte · {g["total_reportes"]} informes</div>
          <h2>{escape(cohort["name"])}</h2>
          <p>Este panel resume cómo se manifiesta el perfil de egreso en los <strong>{g["total_reportes"]} informes</strong> evaluados de esta cohorte. Muestra qué competencias están evidenciadas, cuáles requieren refuerzo y con qué grado de evidencia se acredita cada una.</p>
        </div>
      </section>

      <section class="macro-dashboard-grid">
        <div class="macro-color-card red">
          <span>Porcentaje de logro</span>
          <strong title="Se evaluaron todas las competencias en los {g['total_reportes']} informes: {total_aprobadas} de {total_comps} alcanzaron grado 2 o superior">{aprob_pct:.1f}%</strong>
          <div class="macro-mini-track"><div style="width:{aprob_pct:.1f}%"></div></div>
          <small>Informes que alcanzan al menos grado 2 (Uso concreto) sobre el total evaluado.</small>
        </div>
        <div class="macro-color-card yellow">
          <span>Perfil de egreso cubierto</span>
          <strong title="{cumplidas} de {len(comp_list)} competencias evidenciadas (logro ≥ 70%)">{perfil_pct:.0f}%</strong>
          <div class="macro-mini-track"><div style="width:{perfil_pct:.1f}%"></div></div>
          <small><span class="macro-mini-pill ok"><i class="ok"></i>{cumplidas} evidenciadas</span> <span class="macro-mini-pill mid"><i class="mid"></i>{desarrollo} parcialmente evidenciadas</span> <span class="macro-mini-pill risk"><i class="risk"></i>{brecha} con escasa evidencia</span></small>
        </div>
        <div class="macro-color-card blue">
          <span>Grado de evidencia promedio</span>
          <strong title="Entre todas las competencias evaluadas en los {g['total_reportes']} informes: {g.get('score_actual', 0)} de {g.get('score_max', 0)} puntos acumulados">{g["nivel_promedio_global"]:.2f}/3</strong>
          <div class="macro-mini-track"><div style="width:{nivel_pct:.1f}%;background:{LEVEL_COLORS[_promedio_a_nivel(g['nivel_promedio_global'])]}"></div></div>
          <small>{''.join(f'<span class="macro-mini-pill" style="background:{_hex_to_rgba(LEVEL_COLORS[lvl], 0.12)};color:{LEVEL_COLORS[lvl]}"><i style="background:{LEVEL_COLORS[lvl]};width:9px;height:9px;border-radius:999px;display:inline-block"></i>{lvl}: {LEVEL_LABELS[lvl]}</span> ' for lvl in ['0','1','2','3'])}</small>
        </div>
        <div class="macro-mini-card">
          <span>Competencias por reforzar</span>
          <strong title="{brecha} competencias con menos del 50% de logro, prioritarias para el plan de mejora">{brecha}</strong>
          <p>competencias con menos del 50% de logro. Son las que presentan menor presencia del perfil de egreso en la cohorte.</p>
        </div>
      </section>

      <section class="macro-panel macro-wide-panel">
        <div class="macro-panel-head">
          <div>
            <div class="macro-panel-title">Estado de cada competencia en la cohorte</div>
            <p>Cada tarjeta representa una competencia. El color de la barra lateral y la barra de progreso indican qué tan presente está en los informes de la cohorte.</p>
          </div>
          <div class="macro-status-legend">
            <span title="La mayoría de los informes demuestra esta competencia de forma consistente (logro ≥ 70%)"><i class="ok"></i>Evidenciada (≥ 70%)</span>
            <span title="La competencia aparece en varios informes, pero aún no es consistente (logro entre 50% y 69%)"><i class="mid"></i>Parcialmente evidenciada (50–69%)</span>
            <span title="Existe poca o ninguna evidencia de esta competencia en los informes (logro &lt; 50%)"><i class="risk"></i>Escasa evidencia (&lt; 50%)</span>
          </div>
          <div class="macro-state-descriptions">
            <span><b>Evidenciada</b> — la mayoría de los informes demuestra esta competencia de forma consistente.</span>
            <span><b>Parcialmente evidenciada</b> — la competencia aparece en varios informes, pero aún no es consistente.</span>
            <span><b>Escasa evidencia</b> — existe poca o ninguna evidencia de esta competencia en los informes.</span>
          </div>
        </div>
        <div class="macro-comp-grid">{"".join(tiles)}</div>
      </section>

      <section class="macro-analytics-grid">
        <div class="macro-panel macro-chart-card macro-bars-card">
          <div class="macro-panel-head">
            <div>
              <div class="macro-panel-title">Logro por competencia</div>
              <p>Competencias ordenadas numéricamente. La altura de cada barra indica qué proporción de informes alcanza grado 2 (Uso concreto) o superior en esa competencia.</p>
            </div>
          </div>
          {bar_chart}
        </div>
        <div class="macro-panel macro-chart-card macro-line-panel">
          <div class="macro-panel-head">
            <div>
              <div class="macro-panel-title">Frecuencia vs. profundidad por competencia</div>
                <p>
                  <b>Línea azul (discontinua)</b> = porcentaje de informes donde la competencia se evidencia en grado 2 o superior. Mientras más alta, más frecuente es esa competencia en la cohorte.
                </p>
                <p>
                  <b>Línea roja (continua)</b> = grado de evidencia promedio alcanzado en esa competencia (escala 0 a 3). Mientras más alta, mayor profundidad o dominio demostraron los estudiantes.
                </p>
            </div>
            <div class="macro-chart-legend">
              <span><i class="score"></i>Grado de evidencia (profundidad)</span>
              <span><i class="aprob"></i>Tasa de logro (frecuencia)</span>
            </div>
          </div>
          {line_chart}
          <div class="macro-line-help">
            <span>Ambas líneas altas → competencia evidenciada: se demuestra con frecuencia y profundidad</span>
            <span>Línea azul alta y roja media → aparece en muchos informes, pero con evidencia superficial</span>
            <span>Ambas líneas bajas → escasa evidencia: pocos informes evidencian la competencia</span>
          </div>
        </div>
        <div class="macro-panel macro-distribution-card">
          <div class="macro-panel-title">Distribución del grado de evidencia</div>
          <div class="macro-pie-wrap">
            <div class="macro-ring" style="--n0:{nivel_dist.get("0", 0) / total_comps * 100 if total_comps else 0:.1f};--n1:{nivel_dist.get("1", 0) / total_comps * 100 if total_comps else 0:.1f};--n2:{nivel_dist.get("2", 0) / total_comps * 100 if total_comps else 0:.1f};--n3:{nivel_dist.get("3", 0) / total_comps * 100 if total_comps else 0:.1f}">
              <div><strong title="Total de evaluaciones de todas las competencias en los {g['total_reportes']} informes">{total_comps}</strong><span>evaluaciones</span></div>
            </div>
          </div>
          {_legend_html(nivel_dist, total_comps)}
        </div>
      </section>

      <section class="macro-bottom-grid">
        <div class="macro-panel">
          <div class="macro-panel-title">Competencias evidenciadas</div>
          <p class="macro-panel-note">Competencias con mayor grado de evidencia acumulado. Ordenadas de mayor a menor evidencia.</p>
          {''.join(top_rows) if top_rows else '<p>Sin competencias disponibles.</p>'}
        </div>
        <div class="macro-panel macro-gap-panel">
          <div class="macro-panel-title">Competencias por reforzar</div>
          <p class="macro-panel-note">Competencias con menor porcentaje de logro (grado 2 o superior). Son prioritarias para el plan de mejora.</p>
          {''.join(weak_rows) if weak_rows else '<p>Sin brechas disponibles.</p>'}
        </div>
      </section>

      <section class="macro-panel macro-table-panel">
        <div class="macro-panel-head">
          <div>
            <div class="macro-panel-title">Matriz agregada: detalle por competencia</div>
            <p>Cada fila corresponde a una competencia del perfil de egreso. Los valores consolidan <strong>todos los informes</strong> de la cohorte. El grado de evidencia va de <strong>0 a 3</strong> según lo observado en los informes.</p>
          </div>
          <div class="macro-level-codes">
            <span data-tooltip="No hay mención ni evidencia de la competencia en el informe"><i style="background:{LEVEL_COLORS['0']}"></i><b style="color:{LEVEL_COLORS['0']}">{LEVEL_SHORT["0"]}</b> sin evidencia</span>
            <span data-tooltip="Menciona conceptos o herramientas, pero sin demostrar aplicación práctica"><i style="background:{LEVEL_COLORS['1']}"></i><b style="color:{LEVEL_COLORS['1']}">{LEVEL_SHORT["1"]}</b> solo teoría</span>
            <span data-tooltip="Existe evidencia de aplicación durante la práctica, con resultados descritos"><i style="background:{LEVEL_COLORS['2']}"></i><b style="color:{LEVEL_COLORS['2']}">{LEVEL_SHORT["2"]}</b> uso concreto</span>
            <span data-tooltip="Evidencia sólida: evalúa alternativas, justifica decisiones y reflexiona sobre el impacto"><i style="background:{LEVEL_COLORS['3']}"></i><b style="color:{LEVEL_COLORS['3']}">{LEVEL_SHORT["3"]}</b> dominio técnico</span>
          </div>
        </div>
        <div class="macro-table-scroll">
          <table class="macro-matrix-table">
            <thead>
              <tr>
                <th>Competencia</th>
                <th title="Promedio del grado de evidencia: Sin evidencia / Solo teoría / Uso concreto / Dominio técnico">Nivel</th>
                <th>% Logro</th>
                <th title="Grado 0: Sin evidencia – la competencia no aparece en el informe"><span style="color:{LEVEL_COLORS["0"]}">{LEVEL_SHORT["0"]}</span></th>
                <th title="Grado 1: Solo teoría – se mencionan conceptos pero sin aplicación"><span style="color:{LEVEL_COLORS["1"]}">{LEVEL_SHORT["1"]}</span></th>
                <th title="Grado 2: Uso concreto – evidencia de aplicación durante la práctica"><span style="color:{LEVEL_COLORS["2"]}">{LEVEL_SHORT["2"]}</span></th>
                <th title="Grado 3: Dominio técnico – evidencia sólida con reflexión y justificación"><span style="color:{LEVEL_COLORS["3"]}">{LEVEL_SHORT["3"]}</span></th>
                <th>Clasificación</th>
              </tr>
            </thead>
            <tbody>{"".join(matrix_rows)}</tbody>
          </table>
        </div>
      </section>
    </div>
    """)


def render():
    cohort_id = st.session_state.get("selected_cohort_id")
    cohort = get_cohort(cohort_id) if cohort_id else None

    if not cohort:
        st.warning("No se encontró la cohorte seleccionada.")
        return

    macro = compute_cohort_macro(cohort_id)
    g = macro["global"]
    competencias = macro["competencias"]
    tipo_label = _formatear_tipo(cohort.get("tipo_documento", ""))

    meta_items = [
        badge(tipo_label, "outline"),
        f'{g["total_reportes"]} informe{"s" if g["total_reportes"] != 1 else ""} procesado{"s" if g["total_reportes"] != 1 else ""}',
    ]
    if cohort.get("created_at"):
        meta_items.append(f'Creada: {cohort["created_at"][:10]}')

    breadcrumb = '<a href="#" onclick="alert(\'cohorts\')">Mis Cohortes</a> <span style="color:var(--uandes-text-muted)">/</span> Resultados de cohorte'

    page_hero(
        "Resultados de la cohorte",
        subtitle=f"Evaluación del perfil de egreso en {cohort['name']}: {g['total_reportes']} informe{'s' if g['total_reportes'] != 1 else ''} analizados",
        meta_items=meta_items,
        back_target="cohort_config",
    )

    if g["total_reportes"] == 0:
        st.subheader("Resultados")
        empty_state(
            "Sin resultados",
            "Aún no hay informes procesados in esta cohorte.",
        )
        return

    nivel_dist = {}
    for c in competencias.values():
        for lvl, count in c["distribucion"].items():
            nivel_dist[lvl] = nivel_dist.get(lvl, 0) + count
    total_comps = sum(nivel_dist.values())

    st.markdown(
        _macro_dashboard_html(cohort, tipo_label, g, nivel_dist, total_comps, competencias),
        unsafe_allow_html=True,
    )

    col_a, col_b, col_c, col_d = st.columns(4)
    with col_a:
        if st.button("Resultados Micro", use_container_width=True):
            st.session_state["page"] = "cohort_reports"
            st.rerun()
    with col_b:
        if st.button("Agregar más informes", use_container_width=True):
            st.session_state["new_cohort"] = False
            st.session_state["page"] = "upload"
            st.rerun()
    with col_c:
        st.download_button(
            "Exportar Excel",
            data=_generate_macro_only_excel(cohort.get("report_ids", [])),
            file_name=f"{cohort['name']}_resultados.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"macro_export_{cohort_id}",
            use_container_width=True,
        )
    with col_d:
        st.download_button(
            "Descargar Reporte Tecnico",
            data=_generate_processing_excel(cohort["name"], cohort.get("report_ids", [])),
            file_name=f"Reporte de Procesamiento - {cohort['name']}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"procesamiento_export_{cohort_id}",
            use_container_width=True,
        )
