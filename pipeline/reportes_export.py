import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from pipeline.persistence import load_report


RESUMEN_MACRO_SHEET = "Resumen Macro de Competencias"
RESULTADOS_SHEET = "Resultados de Evaluación"
PROCESAMIENTO_SHEET = "Reporte de Procesamiento"
JPC_APROBACION_MIN = 0.60
TIEMPO_MANUAL_ESTIMADO_MIN = 180


def build_export_index(report_ids: list[str]) -> list[dict]:
    index = []
    for report_id in report_ids:
        report = load_report(report_id)
        if report:
            index.append(report.to_index_entry())
    return index


def _completed_reports(index: list[dict]):
    reports = []
    for entry in sorted(index, key=lambda e: e.get("timestamp", ""), reverse=True):
        report = load_report(entry["report_id"])
        if report and report.estado == "completado":
            reports.append(report)
    return reports


def _trace_by_competencia(report) -> dict[str, dict]:
    trazabilidad = report.reporte_procesamiento.get("trazabilidad_competencias", [])
    return {t.get("competencia_id", ""): t for t in trazabilidad}


def _display_estado_revision(resultado: dict, traza: dict) -> str:
    estado = str(resultado.get("estado_revision", "")).lower()
    nivel = resultado.get("nivel", traza.get("nivel_asignado", 0))
    jpc = traza.get("JPC", 0) or 0
    jpc_aplicable = traza.get("JPC_aplicable", True)

    if "no_aprob" in estado or nivel == 0 or not jpc_aplicable:
        return "No Aprobada"
    if "requiere" in estado or jpc < JPC_APROBACION_MIN:
        return "Requiere Revision"
    return "Respaldo Suficiente"


def _timestamp_validacion(resultado: dict, report) -> str:
    for key in ("timestamp_validacion", "validated_at", "updated_at"):
        if resultado.get(key):
            return str(resultado[key])[:19]
    return str(report.timestamp or datetime.now().isoformat())[:19]


def _resultados_rows(reports) -> tuple[list[str], list[list]]:
    include_report_cols = len(reports) > 1
    headers = [
        "ID",
        "Competencia_Nombre",
        "Nivel",
        "Justificación",
        "Cita_1",
        "Sección_Cita_1",
        "Cita_2",
        "Sección_Cita_2",
        "Cita_3",
        "Sección_Cita_3",
        "JPC_Porcentaje",
        "Estado_Revision",
        "Timestamp_Validación",
    ]
    if include_report_cols:
        headers = ["Reporte", "Report_ID"] + headers

    rows = []
    for report in reports:
        trazas = _trace_by_competencia(report)
        for resultado in report.vista_preliminar:
            cid = resultado.get("competencia_id", "")
            traza = trazas.get(cid, {})
            citas = list(resultado.get("citas", []))[:3]
            secciones = list(resultado.get("secciones_fuente", []))[:3]
            while len(citas) < 3:
                citas.append("")
            while len(secciones) < 3:
                secciones.append("")

            jpc = traza.get("JPC", 0) or 0
            row = [
                cid,
                resultado.get("competencia_nombre", ""),
                resultado.get("nivel", traza.get("nivel_asignado", 0)),
                resultado.get("justificacion", ""),
                citas[0],
                secciones[0],
                citas[1],
                secciones[1],
                citas[2],
                secciones[2],
                jpc,
                _display_estado_revision(resultado, traza),
                _timestamp_validacion(resultado, report),
            ]
            if include_report_cols:
                row = [report.pdf_name, report.report_id] + row
            rows.append(row)

    return headers, rows


def _num(value) -> float:
    return value if isinstance(value, (int, float)) else 0


def _format_minutes(value) -> str:
    if value is None:
        return ""
    return f"{value:g} minutos" if isinstance(value, (int, float)) else str(value)


def _processing_blocks(report) -> list[list]:
    state = report.pipeline_state
    c2 = state.get("c2", {})
    rp = report.reporte_procesamiento
    trazabilidad = rp.get("trazabilidad_competencias", [])
    tiempos = rp.get("tiempos", {})

    detectadas = c2.get("secciones_detectadas", [])
    ausentes = c2.get("secciones_ausentes", [])
    total_encontradas = c2.get("total_secciones", len(detectadas))
    total_esperadas = total_encontradas + len(ausentes)

    total_competencias = len(trazabilidad)
    aprobadas = sum(1 for t in trazabilidad if _num(t.get("JPC")) >= JPC_APROBACION_MIN and t.get("JPC_aplicable", True))
    no_aprobadas = total_competencias - aprobadas
    tasa_aprobacion = aprobadas / total_competencias if total_competencias else 0

    jpc_values = [_num(t.get("JPC")) for t in trazabilidad if t.get("JPC") is not None]
    jpc_promedio = sum(jpc_values) / len(jpc_values) if jpc_values else 0
    max_traza = max(trazabilidad, key=lambda t: _num(t.get("JPC")), default={})
    min_traza = min(trazabilidad, key=lambda t: _num(t.get("JPC")), default={})
    confianza_values = [_num(t.get("F_confianza")) for t in trazabilidad if t.get("F_confianza") is not None]
    similitud_values = [_num(t.get("R_similitud_promedio")) for t in trazabilidad if t.get("R_similitud_promedio") is not None]
    confianza_promedio = sum(confianza_values) / len(confianza_values) if confianza_values else 0
    similitud_promedio = sum(similitud_values) / len(similitud_values) if similitud_values else 0

    t_auto = tiempos.get("T_procesamiento_automatico_min")
    t_revision = tiempos.get("T_revision_humana_min")
    t_ajustes = tiempos.get("T_ajustes_min")
    t_total = tiempos.get("T_IA_total_min")
    if t_total is None:
        t_total = sum(_num(t) for t in [t_auto, t_revision, t_ajustes])
    reduccion = (TIEMPO_MANUAL_ESTIMADO_MIN - t_total) / TIEMPO_MANUAL_ESTIMADO_MIN if t_total else 0

    rows = [
        ["INFORME", report.pdf_name],
        ["Report ID", report.report_id],
        ["Tipo", report.tipo_documento],
        ["Fecha", str(report.timestamp)[:19]],
        [],
        ["ANÁLISIS DE SECCIONES", ""],
        ["Secciones Detectadas", ", ".join(detectadas) if detectadas else ""],
        ["Secciones Ausentes", ", ".join(ausentes) if ausentes else "(ninguna)"],
        ["Total Secciones Esperadas", total_esperadas],
        ["Total Secciones Encontradas", total_encontradas],
        [],
        ["ESTADÍSTICAS DE COBERTURA", ""],
        ["Total Competencias Evaluadas", total_competencias],
        [f"Competencias Aprobadas (JPC >= {JPC_APROBACION_MIN:.2f})", aprobadas],
        [f"Competencias No Aprobadas (JPC < {JPC_APROBACION_MIN:.2f})", no_aprobadas],
        ["Tasa de Aprobación", tasa_aprobacion],
        [],
        ["MÉTRICAS AGREGADAS", ""],
        ["JPC Promedio", round(jpc_promedio, 3)],
        ["JPC Máximo", f"{_num(max_traza.get('JPC')):.3f} ({max_traza.get('competencia_id', '')})" if max_traza else ""],
        ["JPC Mínimo", f"{_num(min_traza.get('JPC')):.3f} ({min_traza.get('competencia_id', '')})" if min_traza else ""],
        ["Confianza Promedio (F)", round(confianza_promedio, 3)],
        ["Similitud Promedio (R)", round(similitud_promedio, 3)],
        [],
        ["TIEMPOS (KPI 2)", ""],
        ["Tiempo Procesamiento Automático", _format_minutes(t_auto)],
        ["Tiempo Revisión Humana (HITL)", _format_minutes(t_revision)],
        ["Tiempo Ajustes", _format_minutes(t_ajustes)],
        ["Tiempo Total con IA", _format_minutes(t_total)],
        ["Tiempo Manual Estimado (Línea Base)", _format_minutes(TIEMPO_MANUAL_ESTIMADO_MIN)],
        ["Reducción de Tiempo (RTR)", reduccion],
        [],
        ["TRAZABILIDAD POR COMPETENCIA", ""],
        [
            "Competencia",
            "Embedding_Generado",
            "Similitud_Calculada",
            "Recuperación_Ejecutada",
            "Chunks_Recuperados",
            "Dictamen_Generado",
            "Estado_Final",
            "JPC_Aplicable",
        ],
    ]
    for traza in trazabilidad:
        rows.append([
            traza.get("competencia_id", ""),
            "✓" if traza.get("embedding_generado") else "",
            "✓" if traza.get("similitud_calculada") else "",
            "✓" if traza.get("recuperacion_ejecutada") else "",
            traza.get("chunks_recuperados", 0),
            "✓" if traza.get("dictamen_generado") else "",
            traza.get("estado_final", ""),
            "✓" if traza.get("JPC_aplicable") else "",
        ])

    ajustes = rp.get("historial_ajustes", [])
    if ajustes:
        rows.extend([[], ["HISTORIAL DE AJUSTES", ""], ["Ajuste_ID", "Competencia", "Solicitud", "Capas_Reprocesadas", "Duración_Min", "Resultado"]])
        for ajuste in ajustes:
            rows.append([
                ajuste.get("ajuste_id", ""),
                ajuste.get("competencia_id", ""),
                ajuste.get("solicitud_usuario", ""),
                ", ".join(ajuste.get("capas_reprocesadas", [])),
                ajuste.get("duracion_min", ""),
                ajuste.get("resultado", ""),
            ])
    return rows


def _write_resultados_sheet(ws, headers: list[str], rows: list[list]):
    ws.append(headers)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="E7E5E4")
    green = PatternFill("solid", fgColor="C6E0B4")
    yellow = PatternFill("solid", fgColor="FFE699")
    red = PatternFill("solid", fgColor="F4CCCC")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    estado_col = headers.index("Estado_Revision") + 1 if "Estado_Revision" in headers else None
    jpc_col = headers.index("JPC_Porcentaje") + 1 if "JPC_Porcentaje" in headers else None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if estado_col:
            estado = row[estado_col - 1].value
            fill = green if estado == "Respaldo Suficiente" else yellow if estado == "Requiere Revision" else red
            row[estado_col - 1].fill = fill
        if jpc_col:
            row[jpc_col - 1].number_format = "0.00%"
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    widths = {
        "ID": 10,
        "Competencia_Nombre": 36,
        "Nivel": 10,
        "Justificación": 48,
        "Cita_1": 48,
        "Sección_Cita_1": 16,
        "Cita_2": 48,
        "Sección_Cita_2": 16,
        "Cita_3": 48,
        "Sección_Cita_3": 16,
        "JPC_Porcentaje": 16,
        "Estado_Revision": 22,
        "Timestamp_Validación": 22,
        "Reporte": 30,
        "Report_ID": 38,
    }
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(header, 18)


def _write_processing_sheet(ws, reports):
    row_idx = 1
    section_fill = PatternFill("solid", fgColor="E7E5E4")
    for report in reports:
        for row in _processing_blocks(report):
            if row:
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    if col_idx == 1 and isinstance(value, str) and value.isupper():
                        cell.font = Font(bold=True)
                        cell.fill = section_fill
                    elif row_idx == 1 or value in ("Competencia", "Ajuste_ID"):
                        cell.font = Font(bold=True)
            row_idx += 1
        row_idx += 2

    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 28 if col_idx == 1 else 24

    for row in ws.iter_rows():
        if row and row[0].value in ("Tasa de Aprobación", "Reducción de Tiempo (RTR)"):
            row[1].number_format = "0.00%"


def _cohort_summary_rows(reports) -> list[list]:
    if not reports:
        return [["No hay informes completados."]]

    all_jpc = []
    all_tiempos = []
    all_comp_counts = []
    total_reportes = len(reports)
    total_errores = 0

    for report in reports:
        rp = report.reporte_procesamiento
        trazabilidad = rp.get("trazabilidad_competencias", [])
        tiempos = rp.get("tiempos", {})

        jpc_values = [_num(t.get("JPC")) for t in trazabilidad if t.get("JPC") is not None]
        all_jpc.extend(jpc_values)
        all_comp_counts.append(len(trazabilidad))

        t_auto = tiempos.get("T_procesamiento_automatico_min")
        t_revision = tiempos.get("T_revision_humana_min")
        t_ajustes = tiempos.get("T_ajustes_min")
        t_total = tiempos.get("T_IA_total_min")
        if t_total is None:
            t_total = sum(_num(t) for t in [t_auto, t_revision, t_ajustes])
        if t_total:
            all_tiempos.append(t_total)

    jpc_promedio_cohorte = round(sum(all_jpc) / len(all_jpc), 4) if all_jpc else 0
    comp_promedio = round(sum(all_comp_counts) / len(all_comp_counts), 1) if all_comp_counts else 0
    tiempo_total_cohorte = round(sum(all_tiempos), 2) if all_tiempos else 0
    jpc_min = round(min(all_jpc), 4) if all_jpc else 0
    jpc_max = round(max(all_jpc), 4) if all_jpc else 0
    comp_min = min(all_comp_counts) if all_comp_counts else 0
    comp_max = max(all_comp_counts) if all_comp_counts else 0
    comp_global = sum(all_comp_counts)

    rows = [
        ["RESUMEN DE COHORTE", ""],
        [],
        ["Información General", ""],
        ["Total Informes Completados", total_reportes],
        ["Total Competencias Evaluadas (global)", comp_global],
        [],
        ["JPC (Índice de Calidad de Procesamiento)", ""],
        ["JPC Promedio Cohorte", jpc_promedio_cohorte],
        ["JPC Mínimo", jpc_min],
        ["JPC Máximo", jpc_max],
        [],
        ["Competencias por Informe", ""],
        ["Promedio Competencias por Informe", comp_promedio],
        ["Mínimo Competencias por Informe", comp_min],
        ["Máximo Competencias por Informe", comp_max],
        [],
        ["Tiempo de Procesamiento", ""],
        ["Tiempo Total Cohorte (min)", _format_minutes(tiempo_total_cohorte)],
        ["Tiempo Promedio por Informe (min)", round(tiempo_total_cohorte / total_reportes, 2) if total_reportes else 0],
        [],
    ]
    for report in reports:
        rows.append([
            report.pdf_name,
            f'{report.timestamp[:19] if hasattr(report, "timestamp") else ""}',
        ])
        rp = report.reporte_procesamiento
        trazabilidad = rp.get("trazabilidad_competencias", [])
        jpc_vals = [_num(t.get("JPC")) for t in trazabilidad if t.get("JPC") is not None]
        tiempos = rp.get("tiempos", {})
        t_total = tiempos.get("T_IA_total_min") or sum(_num(t) for t in [tiempos.get("T_procesamiento_automatico_min"), tiempos.get("T_revision_humana_min"), tiempos.get("T_ajustes_min")])
        rows.append([f"  JPC prom: {round(sum(jpc_vals)/len(jpc_vals),4) if jpc_vals else 0}", f"  Tiempo: {_format_minutes(t_total)}", f"  Comps: {len(trazabilidad)}"])
    return rows


def _compute_macro_data(reports) -> tuple[list[dict], dict]:
    all_rc = []
    for report in reports:
        rp = report.reporte_procesamiento
        trazas = {t.get("competencia_id", ""): t for t in rp.get("trazabilidad_competencias", [])}
        for r in report.vista_preliminar:
            cid = r.get("competencia_id", "")
            traza = trazas.get(cid, {})
            all_rc.append({
                "competencia_id": cid,
                "competencia_nombre": r.get("competencia_nombre", ""),
                "nivel": r.get("nivel", 0),
                "jpc": _num(traza.get("JPC", 0)),
                "c_cobertura": _num(traza.get("C_cobertura_citas", 0)),
                "s_pertinencia": _num(traza.get("S_pertinencia_seccion", 0)),
                "r_similitud": _num(traza.get("R_similitud_promedio", 0)),
                "f_confianza": _num(traza.get("F_confianza", 0)),
                "report_id": report.report_id,
            })

    comp_data: dict[str, dict] = {}
    for rc in all_rc:
        cid = rc["competencia_id"]
        if cid not in comp_data:
            comp_data[cid] = {
                "nombre": rc["competencia_nombre"],
                "niveles": [],
                "jpcs": [],
                "jpcs_all": [],
            }
        comp_data[cid]["niveles"].append(rc["nivel"])
        comp_data[cid]["jpcs_all"].append(rc["jpc"])
        if rc["nivel"] >= 2:
            comp_data[cid]["jpcs"].append(rc["jpc"])

    global_total = len(reports)
    global_niveles = sum(len(d["niveles"]) for d in comp_data.values())
    aprobados = [d for d in all_rc if d["nivel"] >= 2]
    global_aprobadas = len(aprobados)
    global_jpcs = [d["jpc"] for d in aprobados]
    global_cs = [d["c_cobertura"] for d in aprobados]
    global_ss = [d["s_pertinencia"] for d in aprobados]
    global_rs = [d["r_similitud"] for d in aprobados]
    global_fs = [d["f_confianza"] for d in aprobados]

    expected_comps = 0
    comps_per_report = []
    tiempos_auto = []
    for report in reports:
        cs = report.pipeline_state.get("c1", {}).get("competencias_activas", [])
        if cs and expected_comps == 0:
            expected_comps = len(cs)
        prev = report.vista_preliminar
        comps_per_report.append(len(prev))
        t = report.reporte_procesamiento.get("tiempos", {}).get("T_procesamiento_automatico_min")
        if t:
            tiempos_auto.append(t)

    avg_comps = round(sum(comps_per_report) / len(comps_per_report), 1) if comps_per_report else 0
    g = {
        "total_reportes": global_total,
        "total_evaluaciones": global_niveles,
        "competencias_aprobadas": global_aprobadas,
        "c_promedio": round(sum(global_cs) / len(global_cs), 4) if global_cs else 0,
        "s_promedio": round(sum(global_ss) / len(global_ss), 4) if global_ss else 0,
        "r_promedio": round(sum(global_rs) / len(global_rs), 4) if global_rs else 0,
        "f_promedio": round(sum(global_fs) / len(global_fs), 4) if global_fs else 0,
        "jpc_promedio": round(sum(global_jpcs) / len(global_jpcs), 4) if global_jpcs else 0,
        "tiempo_total_min": round(max(tiempos_auto), 2) if tiempos_auto else 0,
        "tiempo_promedio_min": round(sum(tiempos_auto) / len(tiempos_auto), 2) if tiempos_auto else 0,
        "comps_promedio": avg_comps,
        "comps_esperadas": expected_comps,
    }

    return comp_data, g


def _macro_competency_rows(reports) -> list[list]:
    if not reports:
        return [["No hay informes completados."]]

    comp_data, g = _compute_macro_data(reports)

    comps_label = f'{g["comps_promedio"]} de {g["comps_esperadas"]}' if g["comps_esperadas"] else f'{g["comps_promedio"]}'

    rows = [
        ["RESUMEN MACRO DE COMPETENCIAS", ""],
        [],
        ["Total Informes Completados", g["total_reportes"]],
        ["Total Evaluaciones", g["total_evaluaciones"]],
        ["Competencias Aprobadas (N2+N3)", g["competencias_aprobadas"]],
        ["Competencias por Informe (promedio)", comps_label],
        [],
        ["DESGLOSE JPC (solo competencias aprobadas N2+N3)", ""],
        ["C: Cobertura de evidencia textual", g["c_promedio"]],
        ["S: Pertinencia estructural de la sección", g["s_promedio"]],
        ["R: Relevancia semántica de fragmentos", g["r_promedio"]],
        ["F: Confianza del modelo en la clasificación", g["f_promedio"]],
        ["JPC Promedio", g["jpc_promedio"]],
        [],
        ["Tiempo de Procesamiento del Lote (min)", _format_minutes(g["tiempo_total_min"])],
        ["Tiempo Promedio por Informe (min)", _format_minutes(g["tiempo_promedio_min"])],
        [],
        [
            "Competencia",
            "Nombre",
            "Nivel Promedio",
            "Tasa Aprobación",
            "JPC Promedio",
            "Dist. SE (N0)",
            "Dist. NA (N1)",
            "Dist. UC (N2)",
            "Dist. DT (N3)",
            "Estado",
            "Score %",
        ],
    ]

    for cid in sorted(comp_data.keys(), key=lambda x: int("".join(ch for ch in x if ch.isdigit()) or 0) if any(ch.isdigit() for ch in x) else x):
        d = comp_data[cid]
        niveles = d["niveles"]
        jpcs = [j for j in d["jpcs"] if isinstance(j, (int, float))]
        n_total = len(niveles)
        nivel_prom = round(sum(niveles) / n_total, 2) if n_total else 0
        aprobadas = sum(1 for n in niveles if n >= 2)
        tasa_aprob = round(aprobadas / n_total, 4) if n_total else 0
        jpc_prom = round(sum(jpcs) / len(jpcs), 4) if jpcs else 0

        dist = {"0": 0, "1": 0, "2": 0, "3": 0}
        for n in niveles:
            dist[str(int(n))] = dist.get(str(int(n)), 0) + 1

        if tasa_aprob >= 0.70:
            estado = "Consolidada"
        elif tasa_aprob >= 0.50:
            estado = "En desarrollo"
        else:
            estado = "Brecha prioritaria"

        score_pct = round(nivel_prom / 3 * 100, 1) if n_total else 0

        rows.append([
            cid,
            d["nombre"],
            nivel_prom,
            tasa_aprob,
            jpc_prom,
            dist["0"],
            dist["1"],
            dist["2"],
            dist["3"],
            estado,
            f"{score_pct:.1f}%",
        ])

    rows.extend([
        [],
        ["Nota sobre tiempos de procesamiento"],
        [
            "Los tiempos de procesamiento individuales (T_procesamiento_automatico_min) reflejan el tiempo real "
            "transcurrido desde el inicio hasta la finalización de cada informe. Debido a que los informes se "
            "procesan en paralelo con hasta 5 llamadas LLM concurrentes a nivel global, la suma de los tiempos "
            "individuales puede superar el tiempo real de procesamiento del lote completo."
        ],
    ])

    return rows


def _write_macro_sheet(ws, rows: list[list]):
    header_fill = PatternFill("solid", fgColor="E7E5E4")
    green = PatternFill("solid", fgColor="C6E0B4")
    yellow = PatternFill("solid", fgColor="FFE699")
    red = PatternFill("solid", fgColor="F4CCCC")

    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    table_header_row = 18

    for cell in ws[table_header_row]:
        cell.font = Font(bold=True)

    estado_col = None
    if ws.max_row >= table_header_row:
        for idx, cell in enumerate(ws[table_header_row], start=1):
            if cell.value == "Estado":
                estado_col = idx
                break

    for row in ws.iter_rows(min_row=table_header_row + 1, max_row=ws.max_row):
        if estado_col and len(row) >= estado_col:
            val = row[estado_col - 1].value
            if val == "Consolidada":
                row[estado_col - 1].fill = green
            elif val == "En desarrollo":
                row[estado_col - 1].fill = yellow
            elif val == "Brecha prioritaria":
                row[estado_col - 1].fill = red
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A19"
    col_widths = {
        "A": 52, "B": 36, "C": 16, "D": 18, "E": 16,
        "F": 16, "G": 16, "H": 16, "I": 16, "J": 20, "K": 12,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    for row in ws.iter_rows(min_row=2, max_row=table_header_row - 1):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.isupper() and cell.value:
                cell.font = Font(bold=True)


def exportar_excel_multi_hoja(index: list[dict]) -> io.BytesIO:
    output = io.BytesIO()
    wb = Workbook()
    ws_macro = wb.active
    ws_macro.title = RESUMEN_MACRO_SHEET
    ws_resultados = wb.create_sheet(RESULTADOS_SHEET)
    ws_procesamiento = wb.create_sheet(PROCESAMIENTO_SHEET)

    reports = _completed_reports(index)
    if reports:
        macro_rows = _macro_competency_rows(reports)
        _write_macro_sheet(ws_macro, macro_rows)
        headers, rows = _resultados_rows(reports)
        _write_resultados_sheet(ws_resultados, headers, rows)
        _write_processing_sheet(ws_procesamiento, reports)
    else:
        ws_macro.append(["Mensaje", "No hay informes completados disponibles para exportar."])
        _write_resultados_sheet(ws_resultados, ["Mensaje"], [["No hay informes completados disponibles para exportar."]])
        ws_procesamiento.append(["Mensaje", "No hay informes completados disponibles para exportar."])

    wb.save(output)
    output.seek(0)
    return output


def exportar_reporte_individual(report_id: str) -> io.BytesIO:
    report = load_report(report_id)
    if not report:
        return io.BytesIO()
    index = [report.to_index_entry()]
    return exportar_excel_multi_hoja(index)
